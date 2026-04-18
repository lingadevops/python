import openpyxl
wb = openpyxl.load_workbook('transactions.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet.max_row)

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, col)
        print(cell.value)
